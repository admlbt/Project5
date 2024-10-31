from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                           QMessageBox, QFileDialog)
from openpyxl import Workbook
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.teacher import Teacher
from models.subject import Subject
from models.teacher_load import TeacherLoad, LevelType
from views.load_distribution_table import LoadDistributionTable
from utils.load_distributor import LoadDistributor

class LoadTab(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.setup_ui()
        self.load_current_distribution()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Верхняя панель с кнопками
        toolbar = QHBoxLayout()
        
        # Кнопка добавления строки
        add_row_btn = QPushButton("Добавить строку")
        add_row_btn.clicked.connect(self.on_add_row)
        
        # Кнопка автоматического распределения
        auto_distribute_btn = QPushButton("Автоматическое распределение")
        auto_distribute_btn.clicked.connect(self.on_auto_distribute)
        
        # Кнопка сохранения
        save_btn = QPushButton("Сохранить распределение")
        save_btn.clicked.connect(self.on_save_distribution)
        
        # Кнопка экспорта в Excel
        export_btn = QPushButton("Экспорт в Excel")
        export_btn.clicked.connect(self.on_export_to_excel)
        
        # Кнопка сброса
        reset_btn = QPushButton("Сбросить")
        reset_btn.clicked.connect(self.on_reset_distribution)
        
        toolbar.addWidget(add_row_btn)
        toolbar.addWidget(auto_distribute_btn)
        toolbar.addWidget(save_btn)
        toolbar.addWidget(export_btn)
        toolbar.addWidget(reset_btn)
        toolbar.addStretch()
        
        layout.addLayout(toolbar)
        
        # Таблица распределения нагрузки
        self.load_table = LoadDistributionTable(self)
        layout.addWidget(self.load_table)
    
    def load_current_distribution(self):
        """Загрузка текущего распределения из базы данных"""
        try:
            loads = self.main_window.db.query(TeacherLoad).filter(
                TeacherLoad.academic_year == "2023-2024"
            ).all()
            
            for load in loads:
                self.load_table.add_load_row(
                    subject_name=load.subject.name,
                    level=load.level,
                    grade=load.grade,
                    hours=load.hours,
                    teacher_name=load.teacher.name
                )
                
        except Exception as e:
            print(f"Ошибка при загрузке распределения: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке: {str(e)}")
    
    def on_add_row(self):
        """Добавление новой строки в таблицу"""
        self.load_table.add_load_row()
    
    def on_export_to_excel(self):
        """Экспорт распределения в Excel"""
        try:
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить распределение",
                "",
                "Excel files (*.xlsx)"
            )
            
            if file_name:
                wb = Workbook()
                ws = wb.active
                ws.title = "Распределение нагрузки"
                
                # Заголовки
                headers = ["Предмет", "Уровень", "Класс", "Часы", "Учитель"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Данные
                for row in range(self.load_table.rowCount()):
                    ws.cell(row=row+2, column=1, value=self.load_table.item(row, 0).text())
                    ws.cell(row=row+2, column=2, value=self.load_table.cellWidget(row, 1).currentText())
                    ws.cell(row=row+2, column=3, value=self.load_table.cellWidget(row, 2).currentText())
                    ws.cell(row=row+2, column=4, value=self.load_table.cellWidget(row, 3).value())
                    ws.cell(row=row+2, column=5, value=self.load_table.cellWidget(row, 4).currentText())
                
                wb.save(file_name)
                QMessageBox.information(self, "Успех", "Распределение успешно экспортировано")
                
        except Exception as e:
            print(f"Ошибка при экспорте: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте: {str(e)}")

    def on_auto_distribute(self):
        """Автоматическое распределение нагрузки"""
        try:
            reply = QMessageBox.question(
                self, 
                'Подтверждение', 
                'Текущее распределение будет удалено. Продолжить?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, 
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # Получаем всех учителей и предметы
                teachers = self.main_window.db.query(Teacher).all()
                subjects = self.main_window.db.query(Subject).all()
                
                # Создаем экземпляр распределителя нагрузки
                distributor = LoadDistributor(self.main_window.db)
                
                # Очищаем таблицу
                self.load_table.setRowCount(0)
                
                # Для каждого предмета
                for subject in subjects:
                    # Для каждого класса
                    for grade in [10, 11]:
                        # Для каждого уровня
                        for level in [LevelType.BASE.value, LevelType.ADVANCED.value]:
                            hours = subject.hours_10_class if grade == 10 else subject.hours_11_class
                            if hours > 0:
                                # Находим подходящего учителя
                                suitable_teacher = distributor.find_suitable_teacher(
                                    subject, teachers, grade, level
                                )
                                
                                if suitable_teacher:
                                    self.load_table.add_load_row(
                                        subject_name=subject.name,
                                        level=level,
                                        grade=grade,
                                        hours=hours,
                                        teacher_name=suitable_teacher.name
                                    )
                
                QMessageBox.information(self, "Успех", "Нагрузка распределена автоматически")
                
        except Exception as e:
            print(f"Ошибка при автоматическом распред��л��нии: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при распределении: {str(e)}")

    def on_save_distribution(self):
        """Сохранение текущего распределения в базу данных"""
        try:
            # Проверяем наличие конфликтов и превышений нагрузки
            has_problems = False
            for row in range(self.load_table.rowCount()):
                teacher_name = self.load_table.cellWidget(row, 4).currentText()
                teacher = self.main_window.db.query(Teacher).filter(
                    Teacher.name == teacher_name
                ).first()
                
                if teacher:
                    total_hours = 0
                    for r in range(self.load_table.rowCount()):
                        if self.load_table.cellWidget(r, 4).currentText() == teacher_name:
                            total_hours += self.load_table.cellWidget(r, 3).value()
                    
                    if total_hours > teacher.max_hours:
                        has_problems = True
                        break
            
            if has_problems:
                reply = QMessageBox.question(
                    self, 
                    'Предупреждение', 
                    'Обнаружены конфликты или превышение нагрузки. Всё равно сохранить?',
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, 
                    QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.No:
                    return
            
            # Подтверждение
            reply = QMessageBox.question(
                self, 
                'Подтверждение', 
                'Сохранить текущее распределение?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, 
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # Удаляем старые записи
                self.main_window.db.query(TeacherLoad).filter(
                    TeacherLoad.academic_year == "2023-2024"
                ).delete()
                
                # Сохраняем новые записи
                for row in range(self.load_table.rowCount()):
                    subject_name = self.load_table.cellWidget(row, 0).currentText()
                    level = self.load_table.cellWidget(row, 1).currentText()
                    grade = int(self.load_table.cellWidget(row, 2).currentText())
                    hours = self.load_table.cellWidget(row, 3).value()
                    teacher_name = self.load_table.cellWidget(row, 4).currentText()
                    
                    # Получаем объекты учителя и предмета
                    teacher = self.main_window.db.query(Teacher).filter(
                        Teacher.name == teacher_name
                    ).first()
                    subject = self.main_window.db.query(Subject).filter(
                        Subject.name == subject_name
                    ).first()
                    
                    if teacher and subject:
                        load = TeacherLoad(
                            teacher_id=teacher.id,
                            subject_id=subject.id,
                            level=level,
                            grade=grade,
                            hours=hours,
                            academic_year="2023-2024"
                        )
                        self.main_window.db.add(load)
                
                self.main_window.db.commit()
                QMessageBox.information(self, "Успех", "Распределение сохранено")
                
        except Exception as e:
            self.main_window.db.rollback()
            print(f"Ошибка при сохранении: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении: {str(e)}")

    def on_reset_distribution(self):
        """Сброс текущего распределения"""
        try:
            reply = QMessageBox.question(
                self, 
                'Подтверждение', 
                'Вы уверены, что хотите сбросить текущее распределение?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, 
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.load_table.setRowCount(0)
                QMessageBox.information(self, "Успех", "Распределение сброшено")
                
        except Exception as e:
            print(f"Ошибка при сбросе распределения: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Ошибка при сбросе: {str(e)}")